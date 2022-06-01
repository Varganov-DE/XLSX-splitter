#!/usr/bin/env python3
# coding: utf-8 -*-
from tkinter.font import NORMAL
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, Protection
from os.path import join, abspath

class NotAllData(Exception):
    pass

data_path = join('.', "Спецификация ИОС4-ОВ.xlsx")
data_path = abspath(data_path)

wb = load_workbook(filename=data_path, data_only=True, read_only=True)

wsn = list(wb.sheetnames)
print(wsn)

wsdata = None

for i in wsn:
    if wb[i]['D1'].value == 'Маркер': #проверка значения в ячейке D1
        wsdata = i
if wsdata == None:
    raise NotAllData('Нет данных в указанной колонке')

ws = wb[wsdata]
shapka = [cell.value for cell in next(
    ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))]

mandata = {} # Цикл сбора всех данных в колонке, и подсчёт их кол-ва

for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row,
                        max_col=ws.max_column):
    if len(row) > 0:
        marker = row[3].value # Указывает что выборка из 2-й колонки(4-й на листе)
        if marker is not None:
            markerdata = [cell.value for cell in row]
            if marker not in mandata:
                mandata[marker] = []
            mandata[marker].append(markerdata)

for marker in mandata:
    print(f'Маркер {marker}, количество позиций: {len(mandata[marker])}')

wb.close

# Создание отдельной книги xlsx под каждый маркер материалов

for marker in mandata: # Цикл создания новых книг xlsx, по маркеру продукции
    exname, *_ = marker.split()
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    ws.append(shapka) # Добавляем шапку
    for row in mandata[marker]:  # Цикл добавления всех строк с соответствующим маркером
        ws.append(row)

# Тут нужно прописать оформление, сделаю позже

        font = Font(name='Calibri',
                     size=10,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000')
        
        ws.font = font

        fill = PatternFill(fill_type=None,
                     start_color='FFFFFFFF',
                     end_color='FF000000')
        border = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style=None,
                              color='FF000000'),
                     bottom=Side(border_style=None,
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                    color='FF000000')
                    )
        alignment=Alignment(horizontal='general',
                         vertical='bottom',
                         text_rotation=0,
                         wrap_text=False,
                         shrink_to_fit=False,
                         indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                             hidden=False)

    # Конец оформления

    exfilname = join('.', 'Data', (exname + '.xlsx'))
    exfilname = abspath(exfilname)
    print(exfilname)

    wb.save(exfilname)
    wb.close

print('\nДанные по маркеру материалов обработаны')
print('Заявки созданы')