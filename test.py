import openpyxl
import os

wb = openpyxl.load_workbook('Specification_test.xlsx') #работаем с файлом который находится в текущей папке

"""""
os.chdir(path = 'e:\site\coding') #прописываем путь к папке в которой лежит файл, если он в другой папке
wb_new = openpyxl.load_workbook('Specification_test_1.xlsx')

#wb = openpyxl.load_workbook('e:\site\coding\Specification_test_1.xlsx')#делает тоже самое, что и предидущая запись

"""""
wb.sheetnames # имена листов книги
sheet = wb['Книга1'] # конкретный лист
sheet.title # заголовок листа
anotherSheet = wb.active # активный лист
sheet['A1'] # получение ячейки листа
sheet['A1'].value # получение значения ячейки
c = sheet['B1']
c.value # получение значения другой ячейки

# Получение строки, столбца и значения из ячейки:

print('Строка %s, Столбец %s : %s' % (c.row, c.column, c.value))
print('Ячейка %s, : %s' % (c.coordinate, c.value))
print(sheet['C1'].value)

