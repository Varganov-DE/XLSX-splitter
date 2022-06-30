from enum import auto
from appJar import gui
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, Protection
from os.path import join, abspath

# Аргумент №1: data_path - название файла и путь к нему
# Аргумент №2: path_output - название папки и путь к ней
# Аргумент №3: number_cell - координаты ячейки с маркером

def splitter_wb(data_path, path_output, number_cell):

    # задаём параметры работы с файлом:
    wb = load_workbook(filename = data_path, data_only = True, read_only = True)

    wsn = wb.sheetnames # присваивает список листов в книге
    print(f"В файле \"{data_path}\", есть листы: {wsn}.")

    number_cell = ord(number_cell.lower()) - 97 #Вычисляем порядковый номер столбца с маркером:

    # запись в переменную shapka названий всех столбцов:
    ws = wb.active
    shapka = [cell.value for cell in next(
        ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))]

    # Цикл сбора всех данных в промаркированных строках, и подсчёт кол-ва, по каждому маркеру:

    mandata = {} # создаём пустой словарь

    # бежим по всем ячейкам в каждой строке начиная со второй(min_row=2) и с первой колонки(min_col=1), пока в них есть данные(max_row=ws.max_row, max_col=ws.max_column) читаем их значения и записываем в кортеж row: 
    
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row,
                            max_col=ws.max_column): 
        
        if len(row) > 0:
            marker = row[number_cell].value # записываем в переменную значение ячейки в 4-м столбце(3-м по индексу)
            if marker is not None: # если данные есть в ячейке, то:
                markerdata = [cell.value for cell in row] # записываем в переменную список из значений ячеек в строке
                
                if marker not in mandata: # если в словаре нет ещё маркера, то:
                    mandata[marker] = []
                mandata[marker].append(markerdata) # добавляем в словарь к каждому индексу(маркеру), значение(список из значений ячеек в этой строке)

    for marker in mandata: # для каждого индекса(маркера) в словаре:
        print(f'По маркеру {marker}, количество позиций: {len(mandata[marker])}') # выводим длинну кортежа(кол-во строк)

    wb.close # закрываем рабочую книгу

    # прописываем стиль ячеек 


    # Создание отдельной книги xlsx под каждый маркер материалов:

    for marker in mandata: # для каждого индекса словаря
        exname, *_ = marker.split() #преобразует список(list) marker в слово(str) переменная exname, в дальнейшем используется для создания имени файла по имени маркера 
        print(f'\nСоздаём файл по маркеру "{exname}":')
        wb = Workbook() #работаем с новым xlsx файлом (Workbook())
        ws = wb.active #работаем с активным листом
        ws.title = "Заявка" #задаём имя листа

        ws.append(shapka) # Добавляем в лист шапку
        i = 1
        for row in mandata[marker]:  #  для каждого индекса(маркера):
            ws.append(row) # добавляем список, заполняем все строки с соответствующим маркером
            ws.column_dimensions['B'].width = 70
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            
            ws.row_dimensions[i].height = 20
            i += 1


        exfilname = join('.', path_output, ('Заявка ' + exname + '.xlsx')) # прописываем путь и название сохраняемого файла
        exfilname = abspath(exfilname)
        print(exfilname)

        ws.delete_cols(number_cell + 1) # удаляем колонку с маркером

        # сохраняем получившийся файл и переходим к следующему маркеру:

        wb.save(exfilname) # сохраняем файл
        wb.close # закрываем файл



        # копируем стиль ячеек из исходного документа в новый: надо как-то это сделать
    # переходим к следующему маркеру

    print('\nДанные по маркеру материалов обработаны')
    print('Заявки созданы')    

def validate_inputs(data_path, path_output, number_cell):
    
    """ Проверяем, если введенные пользователем значения являются правильными.
 
    Аргументы:
        data_path: выбор входящего файла
        output_dir: выбор папки для хранения новых файлов
        number_cell: буква столбца с маркером
         
    Возвращает:
        True, если ошибка и False, если нет
        Список сообщений об ошибке
    """
    errors = False
    error_msgs = []
 
    # Проверяет, выбран ли XLSX файл
    if Path(data_path).suffix != ".xlsx":
        errors = True
        error_msgs.append("Выберите файл формата XLSX!")

    # Проверяет доступный ли каталог
    if not(Path(path_output)).exists():
        errors = True
        error_msgs.append("Выберите доступную папку")
 
    # Проверяет, выбран ли диапазон
    if len(number_cell) < 1:
        errors = True
        error_msgs.append("Введите название столбца с маркером")
 
    return(errors, error_msgs)

def press(button):
    """ Выполняет нажатие кнопки
 
    Аргументы:
        button: название кнопки. Используем названия Выполнить или Выход
    """
    if button == "Process":
        data_path = app.getEntry("data_path")
        path_output = app.getEntry("path_output")
        number_cell = app.getEntry("number_cell")
        errors, error_msg = validate_inputs(data_path, path_output, number_cell)
        if errors:
            app.errorBox("Error", "\n".join(error_msg), parent=None)
        else:
            splitter_wb(data_path, path_output, number_cell)
    else:
        app.stop()
  
# Создать окно пользовательского интерфейса
app = gui("XLSX Splitter", useTtk=True)
app.setTtkTheme("default")
app.setSize(300, 150)
 
# Добавить интерактивные компоненты
app.addLabel("Выберите XLSX файл")
app.addFileEntry("data_path")

app.addLabel("Выберите папку для сохранения заявок")
app.addDirectoryEntry("path_output")
 
app.addLabel("Буквенное обозначение колонки с маркером")
app.addEntry("number_cell")
 
# Связать кнопки с функцией под названием press
app.addButtons(["Process", "Quit"], press)
 
# Запуск интерфейса
app.go()